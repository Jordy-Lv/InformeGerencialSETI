#!/usr/bin/env python3
"""
Extractor del log de indisponibilidades (`DisponibilidadMensual.xlsx`), la
hoja que el equipo diligencia a mano en SharePoint (Célula 3, compartida entre
Acción Fiduciaria, Bancoldex y EMI) para registrar caídas y si son o no
atribuibles a SETI.

Objetivo: dejar de inferir «atribuible a SETI» solo por el texto de la
categoría del ticket de GLPI (regla de hoy en `cargarGlpi()`, ver
`informe-accion-fiduciaria 1.html`) y cruzarlo contra el dato real que
diligencia el equipo — columna «Atribuible a SETI» (SI/NO/EN ESTUDIO),
enlazada por «NUMERO CASO GLPI».

No reemplaza la clasificación de GLPI: la complementa. Para cada ticket que
`extraer_glpi.py` ya clasificó como «incidente» (no revisión de alerta) en el
periodo, busca su número de caso en la hoja «Indisponibilidades» y reporta qué
dice el equipo. Si no lo encuentra —porque aún no lo diligenciaron—, lo deja
marcado SIN_VERIFICAR: no inventa una atribución donde no hay dato.

No llama a ninguna API: lee el .xlsx real (vía openpyxl — primera dependencia
externa de `automatizacion/`; ver `requirements.txt`) desde
`RUTA_INDISPONIBILIDADES` (ruta directa AL ARCHIVO, no a una carpeta: a
diferencia de RUTA_ONEDRIVE, este archivo no se organiza por mes — es un log
continuo compartido entre varios clientes de SETI) y el CSV que ya dejó
`extraer_glpi.py` en `salida/glpi-<periodo>.csv` para ese mismo periodo.

Uso:

    python3 automatizacion/extraer_indisponibilidades.py                  # el mes que ya cerró
    python3 automatizacion/extraer_indisponibilidades.py --periodo 2026-07
    python3 automatizacion/extraer_indisponibilidades.py --archivo /ruta/a/DisponibilidadMensual.xlsx

Requiere haber corrido antes `extraer_glpi.py` para el mismo periodo (deja
`salida/glpi-<periodo>.csv`); si no está, se avisa y se reporta el log de
indisponibilidades solo, sin cruce.

Desde el 29/07/2026, el HTML SÍ lee esta reconciliación (`cargarGlpi()`, vía
`RECONCILIACION_INDISPONIBILIDADES` en el HTML): un ticket marcado «NO» deja
de contar como incidente atribuible a SETI. «SI», «EN ESTUDIO» o sin match
siguen contando (se mantiene la regla de siempre) — sigue pendiente decidir
con negocio qué hacer específicamente con «EN ESTUDIO». Este script también
corrige `historico_casos.json` para el periodo procesado, para que el número
quede bien desde que se generó, no solo mientras se ve como mes en curso. Ver
docs/2026-07-29-relevo-sesion-28-julio.md §2 y §8.
"""

import argparse
import csv
import io
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from sonda_glpi import cargar_env  # noqa: E402
from insumos_af import adjuntar_historico, archivo_de, cargar_paquete, clasificar_caso_glpi, copiar_resguardo, eliminar_resguardo, escribir_paquete, fijar_periodo, mes_cerrado  # noqa: E402
import historico_casos  # noqa: E402

import os  # noqa: E402

try:
    import openpyxl
except ImportError:
    openpyxl = None

SALIDA = Path(__file__).parent / "salida"
HOJA = "indisponibilidades"
CLIENTE_OBJETIVO = "accion fiduciaria"

# Nombres lógicos -> posibles encabezados en el Excel (normalizados). Buscar
# por coincidencia en vez de posición fija: la hoja puede reordenar columnas
# sin que este script se rompa en silencio.
COLUMNAS_BUSCADAS = {
    "cliente": ["cliente"],
    "caso_glpi": ["numero caso glpi", "numero de caso glpi", "caso glpi"],
    "servicio": ["servicio"],
    "objeto": ["objeto"],
    "atribuible": ["atribuible a seti"],
    "tipo_evento": ["tipo de evento"],
    "inicio": ["fecha hora inicio"],
    "motivo": ["motivo"],
}
COLUMNAS_OBLIGATORIAS = ("cliente", "caso_glpi", "atribuible")


class ErrorIndisponibilidades(Exception):
    pass


# --------------------------------------------------------------------------

def norm(v):
    """Sin acentos, minúsculas, separadores colapsados — misma normalización
    que usan los otros extractores y el propio HTML."""
    s = unicodedata.normalize("NFD", str(v or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9]+", " ", s.lower())
    return s.strip()


def solo_digitos(v):
    return re.sub(r"\D", "", str(v or ""))


def periodo_de(valor):
    """«AAAA-MM» de una celda que puede llegar como datetime/date (lo normal,
    vía openpyxl) o como texto; None si no se puede interpretar."""
    if valor is None:
        return None
    if hasattr(valor, "year") and hasattr(valor, "month"):
        return f"{valor.year:04d}-{valor.month:02d}"
    s = str(valor).strip()
    m = re.match(r"(\d{4})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)  # dd/mm/aaaa, por si acaso
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}"
    return None


def _encontrar_columna(header, candidatos):
    normalizados = [norm(h) for h in header]
    for candidato in candidatos:
        c = norm(candidato)
        for i, h in enumerate(normalizados):
            if c in h:
                return i
    return -1


# --------------------------------------------------------------------------

def leer_indisponibilidades(ruta):
    """Abre el .xlsx y devuelve (filas_cliente, nombre_hoja).

    `filas_cliente` ya viene filtrada a Acción Fiduciaria (descarta de paso
    las filas plantilla de la hoja, que traen "Columna calculada" como
    cliente) y cada fila es un dict por nombre lógico (ver COLUMNAS_BUSCADAS),
    de TODOS los meses — el filtro de periodo lo aplica el llamador.
    """
    if openpyxl is None:
        raise ErrorIndisponibilidades("Falta openpyxl. Instálalo con: pip install -r automatizacion/requirements.txt")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    nombre_hoja = next((s for s in wb.sheetnames if norm(s) == HOJA), None)
    if not nombre_hoja:
        raise ErrorIndisponibilidades(
            f"{ruta} no tiene una hoja «Indisponibilidades». Hojas encontradas: {', '.join(wb.sheetnames)}"
        )
    ws = wb[nombre_hoja]
    filas = list(ws.iter_rows(values_only=True))

    # El encabezado no siempre es la primera fila: puede haber un banner de
    # instrucciones encima (fila combinada, como en la copia real vista el
    # 28/07/2026: "Por favor diligenciar todos los campos..."). Se busca la
    # primera fila con «Cliente» Y «Atribuible a SETI» en vez de asumir un
    # número de fila fijo.
    idx_header = -1
    for i, fila in enumerate(filas[:10]):
        normalizados = [norm(c) for c in fila]
        if "cliente" in normalizados and any("atribuible a seti" in n for n in normalizados):
            idx_header = i
            break
    if idx_header < 0:
        raise ErrorIndisponibilidades(
            f"No se encontró una fila de encabezado con «Cliente» y «Atribuible a SETI» "
            f"en las primeras 10 filas de la hoja «{nombre_hoja}»."
        )

    header = filas[idx_header]
    columnas = {clave: _encontrar_columna(header, candidatos) for clave, candidatos in COLUMNAS_BUSCADAS.items()}
    faltan = [clave for clave in COLUMNAS_OBLIGATORIAS if columnas[clave] < 0]
    if faltan:
        raise ErrorIndisponibilidades(
            f"La hoja «{nombre_hoja}» no tiene las columnas obligatorias: {', '.join(faltan)}. "
            f"Encabezado leído: {list(header)}"
        )

    def valor(fila, clave):
        i = columnas[clave]
        return fila[i] if 0 <= i < len(fila) else None

    filas_cliente = []
    for fila in filas[idx_header + 1:]:
        cliente = valor(fila, "cliente")
        if CLIENTE_OBJETIVO not in norm(cliente):
            continue  # descarta otros clientes de la hoja compartida y las filas plantilla ("Columna calculada")
        filas_cliente.append({clave: valor(fila, clave) for clave in columnas})
    return filas_cliente, nombre_hoja


# --------------------------------------------------------------------------

def leer_incidentes_glpi(ruta_csv, periodo):
    """Lee el CSV que ya dejó extraer_glpi.py para este periodo y devuelve un
    dict {ID_glpi: fila} con los que califican como «incidente» (no revisión)
    — la misma cifra que `inc` en cargarGlpi(). None si el CSV no existe: el
    cruce se omite en vez de tratarlo como "cero incidentes"."""
    if not ruta_csv.exists():
        return None
    texto = ruta_csv.read_text(encoding="utf-8-sig")
    filas = list(csv.DictReader(io.StringIO(texto), delimiter=";"))
    incidentes = {}
    for fila in filas:
        if CLIENTE_OBJETIVO not in norm(fila.get("Entidad", "")):
            continue
        if not str(fila.get("Fecha de apertura", "")).startswith(periodo):
            continue
        if clasificar_caso_glpi(fila.get("Categoría"), fila.get("Tipo")) != "incidente":
            continue
        id_glpi = solo_digitos(fila.get("ID", ""))
        if id_glpi:
            incidentes[id_glpi] = fila
    return incidentes


# --------------------------------------------------------------------------

def reconciliar(filas_indisponibilidades, incidentes_glpi):
    """Para cada incidente de GLPI del periodo, busca su NUMERO CASO GLPI
    entre las filas de indisponibilidades (de cualquier mes: el enlace es por
    ID exacto, no depende de que las fechas coincidan) y reporta qué dice el
    equipo. Sin match: SIN_VERIFICAR — nunca se inventa una atribución."""
    por_caso = {}
    for fila in filas_indisponibilidades:
        digitos = solo_digitos(fila.get("caso_glpi"))
        if digitos:
            por_caso[digitos] = fila

    reconciliadas = []
    for id_glpi, fila_glpi in incidentes_glpi.items():
        match = por_caso.get(id_glpi)
        if match:
            a = norm(match.get("atribuible"))
            if a == "si":
                estado = "SI"
            elif a == "no":
                estado = "NO"
            elif "estudio" in a:
                estado = "EN_ESTUDIO"
            else:
                estado = "SIN_VERIFICAR"  # columna diligenciada con algo que no es SI/NO/EN ESTUDIO
        else:
            estado = "SIN_VERIFICAR"
            match = {}
        reconciliadas.append({
            "id_glpi": id_glpi,
            "categoria_tipo": fila_glpi.get("Categoría") or fila_glpi.get("Tipo", ""),
            "atribuible": estado,
            "servicio": match.get("servicio") or "",
            "objeto": match.get("objeto") or "",
            "tipo_evento": match.get("tipo_evento") or "",
            "motivo": match.get("motivo") or "",
        })
    return reconciliadas


CAMPOS_CSV = ["id_glpi", "categoria_tipo", "atribuible", "servicio", "objeto", "tipo_evento", "motivo"]
CABECERA_CSV = ["ID GLPI", "Categoría/Tipo", "Atribuible a SETI", "Servicio", "Objeto", "Tipo de Evento", "Motivo"]


def a_csv(reconciliadas):
    salida = io.StringIO()
    escritor = csv.writer(salida, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    escritor.writerow(CABECERA_CSV)
    for r in reconciliadas:
        escritor.writerow([r[c] for c in CAMPOS_CSV])
    return salida.getvalue()


def verificar(reconciliadas, pendientes, hubo_cruce):
    """Comprobaciones que deben pasar antes de dar el archivo por bueno."""
    if not hubo_cruce:
        return ["No se cruzó contra GLPI: no se encontró el CSV de esa extracción para este periodo "
                "(corre extraer_glpi.py primero para el mismo --periodo)."]
    problemas = []
    sin_verificar = [r for r in reconciliadas if r["atribuible"] == "SIN_VERIFICAR"]
    if sin_verificar:
        problemas.append(
            f"{len(sin_verificar)} incidente(s) de GLPI sin fila correspondiente en Indisponibilidades "
            "(el equipo aún no diligenció NUMERO CASO GLPI para esos casos, o lo diligenció distinto)."
        )
    en_estudio = [r for r in reconciliadas if r["atribuible"] == "EN_ESTUDIO"]
    if en_estudio:
        problemas.append(
            f"{len(en_estudio)} caso(s) marcados «EN ESTUDIO» — pendiente decidir con negocio cómo "
            "tratarlos (ver docs/2026-07-29-relevo-sesion-28-julio.md §8)."
        )
    if pendientes:
        problemas.append(
            f"{len(pendientes)} indisponibilidad(es) de Acción Fiduciaria en el periodo sin NUMERO CASO "
            "GLPI diligenciado en el Excel."
        )
    return problemas


# --------------------------------------------------------------------------

def main():
    cargar_env()

    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--periodo", default=mes_cerrado(),
                   help="Mes a cruzar, formato AAAA-MM (por defecto, el mes que ya cerró)")
    p.add_argument("--archivo", default=None,
                   help="Ruta al .xlsx de indisponibilidades; por defecto, RUTA_INDISPONIBILIDADES del .env")
    args = p.parse_args()

    ruta_str = args.archivo or os.environ.get("RUTA_INDISPONIBILIDADES", "").strip()
    if not ruta_str:
        # A diferencia de GLPI/AlertOps (credenciales obligatorias), esta fuente es
        # opcional: un equipo donde todavía no exista RUTA_INDISPONIBILIDADES (p. ej.
        # el servidor definitivo, antes de sincronizar esa biblioteca) no debe
        # reportar esto como una falla en actualizar_informe.py — mismo criterio que
        # RUTA_ONEDRIVE en insumos_af.copiar_resguardo().
        print("RUTA_INDISPONIBILIDADES no configurada: se omite el cruce de indisponibilidades "
              "(opcional, no bloqueante). Define la variable en automatizacion/.env o pasa --archivo "
              "para activarlo.")
        return 0
    ruta = Path(ruta_str)
    if not ruta.exists():
        print(f"No se encontró el archivo: {ruta}", file=sys.stderr)
        return 2

    SALIDA.mkdir(exist_ok=True)
    try:
        todas, nombre_hoja = leer_indisponibilidades(ruta)
    except ErrorIndisponibilidades as e:
        print(f"\n{e}", file=sys.stderr)
        return 1

    print(f"Hoja «{nombre_hoja}» de {ruta.name}: {len(todas)} fila(s) de Acción Fiduciaria (todos los meses).")
    del_periodo = [f for f in todas if periodo_de(f.get("inicio")) == args.periodo]
    print(f"  {len(del_periodo)} en {args.periodo}.")

    ruta_glpi = SALIDA / f"glpi-{args.periodo}.csv"
    incidentes_glpi = leer_incidentes_glpi(ruta_glpi, args.periodo)
    hubo_cruce = incidentes_glpi is not None
    if hubo_cruce:
        print(f"GLPI ({ruta_glpi.name}): {len(incidentes_glpi)} incidente(s) atribuible(s) por categoría en {args.periodo}.")
    else:
        print(f"Aviso: no encontré {ruta_glpi} — corre extraer_glpi.py primero para el mismo periodo si quieres el cruce. "
              "Se reporta el log de indisponibilidades solo.", file=sys.stderr)

    reconciliadas = reconciliar(todas, incidentes_glpi or {})
    pendientes = [f for f in del_periodo if not solo_digitos(f.get("caso_glpi"))]
    # Lo que de verdad falta por registrar: incidentes de GLPI sin fila en el
    # Excel de indisponibilidades. Es la señal que decide si el archivo
    # standalone (ver más abajo) debe existir o borrarse.
    pendientes_registro = [r for r in reconciliadas if r["atribuible"] == "SIN_VERIFICAR"]

    # Corrige lo que extraer_glpi.py ya dejó en historico_casos.json: un
    # incidente marcado «NO» por el equipo deja de contar como atribuible a
    # SETI. «SI», «EN ESTUDIO» o SIN_VERIFICAR se mantienen (regla de
    # siempre) — mismo criterio que aplica el HTML en vivo (cargarGlpi(), vía
    # RECONCILIACION_INDISPONIBILIDADES). Sin cruce (hubo_cruce=False) no hay
    # nada que corregir: el valor de extraer_glpi.py queda tal cual.
    datos_historico = None
    if hubo_cruce:
        incidentes_corregidos = sum(1 for r in reconciliadas if r["atribuible"] != "NO")
        datos_historico = historico_casos.cargar()
        anterior = datos_historico["periodos"].get(args.periodo, {}).get("incidentes")
        historico_casos.actualizar_periodo(datos_historico, args.periodo, incidentes=incidentes_corregidos)
        historico_casos.escribir(datos_historico)
        if anterior is not None and anterior != incidentes_corregidos:
            print(f"Histórico corregido: incidentes de {args.periodo} pasó de {anterior} a "
                  f"{incidentes_corregidos} (excluidos los marcados «NO» en indisponibilidades).")

    contenido = a_csv(reconciliadas)
    bytes_csv = contenido.encode("utf-8-sig")
    nombre = f"indisponibilidades-{args.periodo}.csv"
    destino = SALIDA / nombre

    # El archivo standalone (local + copia en RUTA_ONEDRIVE) es una ALERTA:
    # su sola existencia dice «hay un incidente de GLPI sin registrar en el
    # Excel de indisponibilidades». Una vez el equipo registra todos los
    # casos del periodo (SI/NO/EN ESTUDIO, ya no SIN_VERIFICAR), deja de
    # hacer falta — se borra en vez de quedarse ahí con un «sin
    # observaciones» que ya no aporta nada.
    if hubo_cruce and not pendientes_registro:
        if destino.exists():
            destino.unlink()
            print(f"Nada pendiente por registrar: se eliminó {destino} (ya no es necesario).")
        borrado = eliminar_resguardo(nombre, args.periodo)
        if borrado:
            print(f"También se eliminó la copia en OneDrive: {borrado}")
        resguardo = None
    else:
        destino.write_bytes(bytes_csv)
        # proteger=False: a diferencia de glpi-*.csv/alertops-*.csv (una foto
        # fija de lo que dijo la fuente al momento de extraer), este archivo
        # cambia legítimamente cada vez que el equipo llena más filas del
        # Excel — protegerlo impediría que una corrida futura actualizara un
        # SIN_VERIFICAR a SI/NO.
        resguardo = copiar_resguardo(destino, nombre, args.periodo, proteger=False)

    js = SALIDA / "insumos-af.js"
    paquete = cargar_paquete(js)
    discrepancia = fijar_periodo(paquete, args.periodo)
    if discrepancia:
        print(f"Aviso: el insumo ya traía periodo {discrepancia} (de otra fuente); "
              f"se ajusta a {args.periodo}.", file=sys.stderr)
    # El insumo para el HTML SIEMPRE lleva la reconciliación completa mientras
    # hubo cruce — exista o no el archivo standalone en disco. Es lo que
    # mantiene correcto el conteo de «atribuible a SETI» en vivo aunque todo
    # ya esté registrado y el archivo visible haya desaparecido.
    if hubo_cruce:
        paquete["archivos"]["indisponibilidades"] = archivo_de(bytes_csv, nombre, f"{ruta.name} · hoja {nombre_hoja}")
    else:
        paquete["archivos"].pop("indisponibilidades", None)
    if datos_historico is not None:
        adjuntar_historico(paquete, datos_historico)
    escribir_paquete(js, paquete)

    problemas = verificar(reconciliadas, pendientes, hubo_cruce)
    print(f"\n{len(reconciliadas)} incidente(s) reconciliado(s).")
    print(f"Archivo standalone: {destino if destino.exists() else '(no existe: nada pendiente por registrar)'}")
    print(f"Copia de resguardo: {resguardo if resguardo else '(sin copia — RUTA_ONEDRIVE no configurada, o nada pendiente)'}")
    print(f"Insumo para el informe: {js}")
    if hubo_cruce:
        print(f"  sha256 {paquete['archivos']['indisponibilidades']['sha256'][:16]}…")
        print("  El HTML ya usa esta reconciliación: un incidente marcado «NO» deja de contar como "
              "atribuible a SETI (ver RECONCILIACION_INDISPONIBILIDADES en informe-accion-fiduciaria 1.html).")
    if problemas:
        print("\n  Revisar antes de usarlo:")
        for x in problemas:
            print(f"    · {x}")
    else:
        print("\n  Verificación: sin observaciones.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
