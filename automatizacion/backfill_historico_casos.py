#!/usr/bin/env python3
"""
Backfill de una sola vez: puebla `historico_casos.json` con los meses ya
cerrados de la hoja «Casos» del consolidado, de antes de que existiera esta
automatización. No pisa ningún periodo que el ledger ya tenga (por ejemplo,
si `extraer_glpi.py`/`extraer_alertas.py` ya corrieron en vivo para
julio-2026): usa upsert «solo si falta» (`historico_casos.solo_si_falta`).

Motivo: el histórico de casos del HTML (el modal de evolución en slide 5,
`DATA_CASOS.historico`) dependía por completo de esta hoja del Excel — había
que cargar el consolidado a mano para que el informe supiera algo de meses
anteriores. Con este backfill (una vez) + la automatización mensual (de aquí
en adelante), deja de depender de ella. Ver
docs/2026-07-29-relevo-sesion-28-julio.md §5.1 y §8.

Uso:

    python3 automatizacion/backfill_historico_casos.py --archivo "/ruta/Disponibilidad Consolidado Mayo.xlsx"
    python3 automatizacion/backfill_historico_casos.py --archivo ... --desde 2025-09 --hasta 2026-06

Por defecto, `--desde` es septiembre de 2025 (inicio de contrato, mismo valor
por defecto que usa el HTML — `new Date(2025,8,1)` en cargarCasos()) y
`--hasta` es «todos los periodos que traiga el Excel» — la protección real
contra pisar datos en vivo no es el rango de fechas, es `solo_si_falta`.
"""

import argparse
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import historico_casos  # noqa: E402

try:
    import openpyxl
except ImportError:
    openpyxl = None

R_ALERTAS = re.compile(r"^alertas")
R_REQUERIMIENTOS = re.compile(r"^requerim")
R_INCIDENTES = re.compile(r"^incidentes?$")


class ErrorBackfill(Exception):
    pass


def norm(v):
    s = unicodedata.normalize("NFD", str(v or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9]+", " ", s.lower())
    return s.strip()


def numero(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def leer_casos_excel(ruta):
    """Devuelve {«AAAA-MM»: {alertas, requerimientos, incidentes}} a partir de
    la hoja «Casos» del consolidado — misma heurística de encabezado y misma
    clasificación de filas que `cargarCasos()` en el HTML."""
    if openpyxl is None:
        raise ErrorBackfill("Falta openpyxl. Instálalo con: pip install -r automatizacion/requirements.txt")

    wb = openpyxl.load_workbook(ruta, data_only=True)
    if "Casos" not in wb.sheetnames:
        raise ErrorBackfill(f"{ruta} no tiene una hoja «Casos». Hojas encontradas: {', '.join(wb.sheetnames)}")
    filas = list(wb["Casos"].iter_rows(values_only=True))

    # La fila de encabezado es la que tiene más columnas con fecha — evita el
    # bloque superior de categorías (Bloqueos, Conexiones...), que no lleva
    # fechas. Mismo criterio que cargarCasos() en el HTML.
    idx_header, mejor = -1, 2
    for i, fila in enumerate(filas):
        n_fechas = sum(1 for v in fila if hasattr(v, "year") and hasattr(v, "month"))
        if n_fechas > mejor:
            mejor, idx_header = n_fechas, i
    if idx_header < 0:
        raise ErrorBackfill("No se encontró una fila de encabezado con fechas en la hoja «Casos».")

    header = filas[idx_header]
    columnas = [(j, v.year, v.month) for j, v in enumerate(header) if hasattr(v, "year") and hasattr(v, "month")]
    if not columnas:
        raise ErrorBackfill("La fila de encabezado de «Casos» no tiene columnas de fecha.")

    def fila_de(patron):
        for fila in filas[idx_header + 1:]:
            etiqueta = next((v for v in fila if v is not None and str(v).strip() != ""
                              and not (hasattr(v, "year") and hasattr(v, "month"))), None)
            if etiqueta is not None and patron.search(norm(etiqueta)):
                return fila
        return None

    fila_alertas = fila_de(R_ALERTAS)
    fila_requerimientos = fila_de(R_REQUERIMIENTOS)
    fila_incidentes = fila_de(R_INCIDENTES)
    if not any([fila_alertas, fila_requerimientos, fila_incidentes]):
        raise ErrorBackfill("La hoja «Casos» no tiene filas Alertas/Requerimientos/Incidentes reconocibles.")

    por_periodo = {}
    for j, anio, mes in columnas:
        clave = f"{anio:04d}-{mes:02d}"
        valores = {
            "alertas": numero(fila_alertas[j]) if fila_alertas and j < len(fila_alertas) else None,
            "requerimientos": numero(fila_requerimientos[j]) if fila_requerimientos and j < len(fila_requerimientos) else None,
            "incidentes": numero(fila_incidentes[j]) if fila_incidentes and j < len(fila_incidentes) else None,
        }
        por_periodo[clave] = valores
    return por_periodo


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--archivo", required=True, help="Ruta al Excel consolidado (con hoja «Casos»)")
    p.add_argument("--desde", default="2025-09",
                   help="Primer periodo a incluir, AAAA-MM (por defecto, inicio de contrato: 2025-09)")
    p.add_argument("--hasta", default=None,
                   help="Último periodo a incluir, AAAA-MM (por defecto, todos los que traiga el Excel)")
    args = p.parse_args()

    ruta = Path(args.archivo)
    if not ruta.exists():
        print(f"No se encontró el archivo: {ruta}", file=sys.stderr)
        return 2

    try:
        por_periodo = leer_casos_excel(ruta)
    except ErrorBackfill as e:
        print(f"\n{e}", file=sys.stderr)
        return 1

    datos = historico_casos.cargar()
    escritos, omitidos, vacios = [], [], []
    for clave in sorted(por_periodo):
        if args.desde and clave < args.desde:
            continue
        if args.hasta and clave > args.hasta:
            continue
        campos = {k: v for k, v in por_periodo[clave].items() if v is not None}
        if not campos:
            vacios.append(clave)
            continue
        if historico_casos.solo_si_falta(datos, clave, **campos):
            escritos.append(clave)
        else:
            omitidos.append(clave)

    historico_casos.escribir(datos)
    print(f"Backfill desde {ruta.name}, hoja «Casos» (rango {args.desde or '(inicio)'} a {args.hasta or '(fin)'}).")
    print(f"  Escritos ({len(escritos)}): {', '.join(escritos) or '(ninguno)'}")
    print(f"  Ya existían en el ledger, no se tocaron ({len(omitidos)}): {', '.join(omitidos) or '(ninguno)'}")
    if vacios:
        print(f"  Sin ningún valor en el Excel para esos meses, se omitieron ({len(vacios)}): {', '.join(vacios)}")
    print(f"\nLedger: {historico_casos.RUTA}")
    print("Corre extraer_glpi.py/extraer_alertas.py (o actualizar_informe.py) para que este backfill "
          "quede también dentro de insumos-af.js y llegue al HTML.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
